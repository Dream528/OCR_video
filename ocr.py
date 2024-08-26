import cv2
import pytesseract
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

def clean_number_strings(text_array):
    # Initialize an empty list to store cleaned numbers
    cleaned_numbers = []
    
    # Iterate over each text string in the array
    for text in text_array:
         # Remove trailing commas
        cleaned_text = text.rstrip(',')
        # Add the cleaned text to the list
        cleaned_numbers.append(cleaned_text)
    
    return list(filter(lambda cleaned_number: cleaned_number.strip() != "", cleaned_numbers))

# Function to extract text from an image using Tesseract OCR
def extract_text_from_frame(frame):
    # Convert frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Use adaptive thresholding to enhance the text
    # This can help in making the text more distinct
    _,enhanced = cv2.threshold(gray,17, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    text = pytesseract.image_to_string(enhanced, config='--psm 6 --oem 3 -c tessedit_char_unblacklist=0123456789')
    textArray = text.split('\n')

    return clean_number_strings(textArray)

def extract_white_text_from_frame(frame):
    # Convert frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    _, textArray = cv2.threshold(gray, 100, 255, cv2.THRESH_BINARY_INV)
    # Use Tesseract to extract text
    text = pytesseract.image_to_string(textArray)
    textArray = text.split('\n')
    return clean_number_strings(textArray)

# Function to crop a specific region from a frame
def crop_frame(frame, x, y, width, height):
    return frame[y:y+height, x:x+width]

# Function to process video and extract data
def process_video(video_path, crop_b_bidprice, crop_b_orders, crop_b_qty, crop_r_offer, crop_r_orders, crop_r_qty, crop_extra):
    # Open the video file
    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    interval = int(fps)  # 1 second interval

    data = []
    last_llt = None

    frame_count = 0

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        # Get the current frame position
        frame_id = int(cap.get(cv2.CAP_PROP_POS_FRAMES))

        # Process every 1 second
        if frame_id % interval == 0:
            # Crop the frame to the desired region
            x1, y1, width1, height1 = crop_b_bidprice
            x2, y2, width2, height2 = crop_b_orders
            x3, y3, width3, height3 = crop_b_qty
            x4, y4, width4, height4 = crop_r_offer
            x5, y5, width5, height5 = crop_r_orders
            x6, y6, width6, height6 = crop_r_qty

            x, y, width, height = crop_extra
            frame_b_bidprice = crop_frame(frame, x1, y1, width1, height1)
            frame_b_orders = crop_frame(frame, x2, y2, width2, height2)
            frame_b_qty = crop_frame(frame, x3, y3, width3, height3)
            frame_r_offer = crop_frame(frame, x4, y4, width4, height4)
            frame_r_orders = crop_frame(frame, x5, y5, width5, height5)
            frame_r_qty = crop_frame(frame, x6, y6, width6, height6)
            frame_extra = crop_frame(frame, x, y, width, height)

            cv2.imwrite("./images/frame_b_bidprice"+ frame_count +".png", frame_b_bidprice)
            frame_count += 1

            # Extract text from the cropped frame
            text_b_bidprice = extract_text_from_frame(frame_b_bidprice)
            text_b_orders = extract_text_from_frame(frame_b_orders)
            text_b_qty = extract_text_from_frame(frame_b_qty)
            text_r_offer = extract_text_from_frame(frame_r_offer)
            text_r_orders = extract_text_from_frame(frame_r_orders)
            text_r_qty = extract_text_from_frame(frame_r_qty)
            text_extra = extract_white_text_from_frame(frame_extra)

            cv2.image
            values = []
            for i in range(0, 20):
                value = {}
                value['Bid Price'] = text_b_bidprice[i].replace(',', '')
                value['Orders'] = text_b_orders[i]
                value['QTY'] = text_b_qty[i]
                value['Offer'] = text_r_offer[i]
                value['Orders1'] = text_r_orders[i]
                value['QTY1'] = text_r_qty[i]
                values.append(value)
            values[0]['Open'] = text_extra[0]
            values[0]['High'] = text_extra[5]
            values[0]['Low'] = text_extra[1]
            values[0]['Prev.Close'] = text_extra[6]
            values[0]['Volumn'] = text_extra[2].replace(',', '')
            values[0]['Avg.Price'] = text_extra[7]
            values[0]['Lower circuit'] = text_extra[4]
            values[0]['Upper circuit'] = text_extra[9]
            values[0]['LTQ'] = text_extra[3]
            values[0]['LTT'] = text_extra[8]
            data = values + data
    cap.release()
    return data

# Function to update Excel file
def update_excel(file_path, new_data):
    number_columns = ['Bid Price', 'QTY', 'Open', 'Low', 'LTQ', 'Lower circuit', 'High', 'Prev.Close', 'Volumn', 'Avg.Price', 'Upper circuit']
    text_columns = ['Orders', 'Offer', 'Orders1', 'QTY1']
    date_columns = ['LTT']
    column_widths = {0: 12, 1: 12, 2: 12, 3: 12, 4: 12, 5: 12, 6: 17, 7: 17, 8: 20, 9: 17, 10: 17, 11: 17, 12: 17, 13: 17, 14: 20, 15: 20, 16: 30}
    # Check if the file exists
    if not os.path.exists(file_path):
        # Create a new Excel file if it doesn't exist
        wb = Workbook()
        wb.save(file_path)

    # Convert new data to DataFrame
    new_data_df = pd.DataFrame(new_data)

    # Load existing data
    try:
        existing_data = pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        existing_data = pd.DataFrame()

    # Concatenate new data on top of existing data
    updated_data = pd.concat([new_data_df, existing_data], ignore_index=True)

    # Convert specified columns to numbers
    if number_columns:
        for col in number_columns:
            updated_data[col] = pd.to_numeric(updated_data[col], errors='coerce')

    # Convert specified columns to dates
    if date_columns:
        for col in date_columns:
            updated_data[col] = pd.to_datetime(updated_data[col], errors='coerce')

    # Convert specified columns to text
    if text_columns:
        for col in text_columns:
            updated_data[col] = updated_data[col].astype(str)

    # Save updated data to Excel
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
        updated_data.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Set column widths if specified
        if column_widths:
            for col_idx, width in column_widths.items():
                col_letter = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = width

# Main function
def main():
    video_path = 'resource.mp4'
    excel_path = 'output.xlsx'  # Ensure this path is correct

    # Define the region of interest (x, y, width, height)
    crop_b_bidprice = (0, 170, 80, 490)
    crop_b_orders = (80, 170, 43, 490)
    crop_b_qty = (123, 170, 52, 490)
    crop_r_offer = (175, 170, 60, 490)
    crop_r_orders = (235, 170, 55, 490)
    crop_r_qty = (290, 170, 60, 490)
    crop_extra = (0, 750, 400, 500)

    # Process video and extract data
    new_data = process_video(video_path, crop_b_bidprice, crop_b_orders, crop_b_qty, crop_r_offer, crop_r_orders, crop_r_qty, crop_extra)

    # Update Excel file with new data
    update_excel(excel_path, new_data)

if __name__ == "__main__":
    main()